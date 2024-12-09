# %%
import os
import sys
import platform
import time
import subprocess
import pandas as pd
from openpyxl import load_workbook
from zipfile import BadZipFile

# Intentar importar IPython.display para utilizar display si está disponible
try:
    from IPython.display import display
    ipython_available = True
except ImportError:
    ipython_available = False

# Intentar importar clear_output de IPython.display para limpiar la salida en Jupyter
try:
    from IPython.display import clear_output
    in_jupyter = True
except ImportError:
    in_jupyter = False


class Utilidades:
    @staticmethod
    def limpiar_consola():
        # Limpia la consola dependiendo del sistema operativo.
        # Si el sistema es Windows, usa el comando 'cls'; si es Linux o MacOS (Darwin), usa 'clear'.
        # Para otros sistemas, no realiza ninguna acción.
        sistema = platform.system()
        os.system('cls' if sistema == 'Windows' else 'clear' if sistema in ['Linux', 'Darwin'] else '')

    @staticmethod
    def cuenta_regresiva(gestor_directorios):
        # Pregunta al usuario si desea salir
        if input('¿Está seguro de que desea salir? (y/n): ').strip().lower() == 'y':
            # Restaurar el directorio original
            os.chdir(gestor_directorios.directorio_original)
            print(f'Directorio restaurado a: {gestor_directorios.directorio_original}')

            # Cuenta regresiva antes de salir
            for i in range(5, 0, -1):
                if in_jupyter:
                    clear_output(wait=True)
                    print(f'Saliendo en {i}...', end='\r')
                else:
                    print(f'Saliendo en {i}...', end='\r')
                    sys.stdout.flush()
                time.sleep(1)


class UtilidadesArchivos:
    @staticmethod
    def listar_archivos_xlsx(directorio, gestor_directorios=None):
        # Lista todos los archivos .xlsx en el directorio especificado
        archivos = [f for f in os.listdir(directorio) if f.endswith('.xlsx')]
        
        # Si no se encuentran archivos .xlsx, solicita al usuario seleccionar otro directorio
        if not archivos:
            print('No se encontraron archivos .xlsx en el directorio especificado.')
            input('Pulse Enter para seleccionar otro directorio...')
            
            # Si se proporciona un gestor de directorios, permite seleccionar un nuevo directorio
            if gestor_directorios:
                directorio = gestor_directorios.obtener_directorio_trabajo()
                return UtilidadesArchivos.listar_archivos_xlsx(directorio, gestor_directorios)
            else:
                return {}  # Devuelve un diccionario vacío si no hay archivos .xlsx y no hay gestor de directorios
        
        # Crea un diccionario numerado de los archivos encontrados
        archivos_dict = {i + 1: archivo for i, archivo in enumerate(archivos)}
        print('\nArchivos .xlsx encontrados:')
        
        # Muestra los archivos enumerados en la consola
        for num, archivo in archivos_dict.items():
            print(f'{num}: {archivo}')
        return archivos_dict

    @staticmethod
    def cuenta_regresiva(gestor_directorios):
        """
        Realiza una cuenta regresiva antes de salir, restaurando el directorio original.
        """
        if input('¿Está seguro de que desea salir? (y/n): ').strip().lower() == 'y':
            os.chdir(gestor_directorios.directorio_original)
            print(f'Directorio restaurado a: {gestor_directorios.directorio_original}')

            for i in range(5, 0, -1):
                if in_jupyter:
                    clear_output(wait=True)
                    print(f'Saliendo en {i}...', end='\r')
                else:
                    print(f'Saliendo en {i}...', end='\r')
                    sys.stdout.flush()
                time.sleep(1)


class GestorDirectorios:
    def __init__(self, in_colab, directorio_externo=None):
        """
        Inicializa la clase de gestión de directorios.
        
        Parámetros:
        - in_colab (bool): Indica si el entorno es Google Colab.
        - directorio_externo (str o None): Ruta proporcionada por otro programa si es llamado.
        """
        self.in_colab = in_colab

        # Usar el directorio externo si está definido
        if directorio_externo and os.path.exists(directorio_externo):
            self.directorio_base = directorio_externo
        else:
            # Si no hay directorio externo, verificar si el programa tiene un directorio actual establecido
            self.directorio_base = os.getcwd()

            # Solo montar el entorno si no se ha definido un directorio en ejecución
            if in_colab and self.directorio_base == '/content':
                # Montar Google Drive y establecer el directorio base predeterminado
                print("Montando Google Drive en entorno Colab.")
                from google.colab import drive
                drive.mount('/content/drive')
                self.directorio_base = '/content/drive/My Drive'

        # Establecer el directorio actual como el directorio base
        self.directorio = self.directorio_base
        self.directorio_original = self.directorio  # Guardar el directorio original

        print(f'Directorio de trabajo inicial: {self.directorio}')

    def formatear_ruta(self, ruta, contexto=''):
        """
        Formatea las rutas para que sean compatibles con el sistema operativo actual.
        """
        ruta = ruta.replace('\\', os.path.sep).replace('/', os.path.sep)
        sistema = platform.system()

        if contexto == 'salida':
            return ruta.replace('/', '\\') if sistema == 'Windows' else ruta.replace('\\', '/')
        else:
            return ruta.replace('/', '\\\\') if sistema == 'Windows' else ruta

    def obtener_directorio_trabajo(self):
        """
        Solicita un nuevo directorio de trabajo al usuario y actualiza la ruta actual.
        """
        # Guarda el directorio actual antes de cambiar
        self.directorio_original = self.directorio

        # Solicita el nuevo directorio al usuario
        nuevo_directorio = input('Introduzca la ruta de trabajo: ').strip()
        nuevo_directorio = self.formatear_ruta(nuevo_directorio)

        if os.path.exists(nuevo_directorio):
            self.directorio = nuevo_directorio
            os.chdir(self.directorio)
            print(f'Directorio cambiado a: {self.directorio}')
        else:
            print(f'Error: El directorio {nuevo_directorio} no existe. Se mantiene el directorio actual.')
        
        return self.directorio


class VisualizadorDataframes:
    def ajustar_ancho_pantalla(self, df):
        # Ajusta la configuración de pantalla para mostrar el DataFrame completo en la consola
        # Establece sin límite el número de filas, columnas y el ancho máximo de cada columna
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.max_colwidth', None)

        # Calcula el ancho total aproximado en caracteres
        ancho_total = sum(len(str(col)) for col in df.columns) + df.shape[1] * 10
        ancho_pantalla = pd.get_option('display.width')

        # Ajusta el ancho de pantalla solo si el DataFrame es más ancho que el límite actual
        if ancho_total > ancho_pantalla:
            pd.set_option('display.width', ancho_total)
    
    def resetear_ancho_pantalla(self):
        # Restablece las opciones de visualización a sus valores predeterminados
        pd.reset_option('display.width')
        pd.reset_option('display.max_columns')
        pd.reset_option('display.max_rows')
        pd.reset_option('display.max_colwidth')

    def visualizar_dataframes(self, directorio, archivos):
        # Carga y visualiza cada hoja de los archivos Excel seleccionados
        for archivo_idx, archivo in enumerate(archivos):
            ruta_archivo = os.path.join(directorio, archivo)
            
            try:
                # Intenta cargar el archivo de Excel
                xls = pd.ExcelFile(ruta_archivo)
                print(f'\nArchivo: {archivo}')
                
                # Itera a través de cada hoja del archivo
                for hoja_idx, nombre_hoja in enumerate(xls.sheet_names):
                    # Lee la hoja actual en un DataFrame
                    df = pd.read_excel(xls, nombre_hoja)
                    print(f'\n--- Hoja: {nombre_hoja} ---')
                    
                    # Ajusta el ancho de pantalla antes de mostrar el DataFrame
                    self.ajustar_ancho_pantalla(df)
                    display(df)  # Muestra el DataFrame en pantalla
                    self.resetear_ancho_pantalla()  # Restablece el ancho de pantalla después de mostrar

                    # Verifica si es la última hoja y el último archivo
                    es_ultima_hoja = (hoja_idx == len(xls.sheet_names) - 1)
                    es_ultimo_archivo = (archivo_idx == len(archivos) - 1)
                    
                    # Solicita al usuario que continúe o regrese al menú
                    if es_ultima_hoja and es_ultimo_archivo:
                        input('\nÚltima hoja. Pulsar Enter para volver.')
                    else:
                        respuesta = input('\nPresione Enter para avanzar o escriba "m" para regresar a menu: ').strip().lower()
                        if respuesta == 'm':
                            return  # Salir al menú si el usuario elige "m"

            # Captura y muestra cualquier error al cargar el archivo
            except Exception as e:
                print(f'Error al cargar {archivo}: {e}')


class ProcesadorArchivosExcel:
    def es_archivo_excel_valido(self, ruta_archivo):
        # Verifica si un archivo es válido para abrirlo con openpyxl.
        # Intenta abrir el archivo en modo binario ('rb') y cargarlo como libro de trabajo de Excel.
        # Si se produce una excepción BadZipFile, KeyError o ValueError, el archivo no es válido.
        try:
            with open(ruta_archivo, 'rb') as file:
                load_workbook(filename=file)
            return True
        except (BadZipFile, KeyError, ValueError):
            return False

    def limpiar_hoja_actualizada(self, df):
        # Elimina columnas y filas vacías, y filtra filas por celdas no vacías sin columna adicional.
        # Primero, elimina todas las columnas y luego las filas que están completamente vacías.
        df = df.dropna(axis=1, how='all')
        df = df.dropna(axis=0, how='all')
        
        # Añade una columna 'non_empty_count' para contar las celdas no vacías en cada fila.
        # Filtra para mantener solo las filas con más de dos celdas no vacías.
        df['non_empty_count'] = df.apply(lambda row: row.count(), axis=1)
        df = df[df['non_empty_count'] > 2]
        
        # Elimina la columna auxiliar 'non_empty_count' y restablece el índice.
        df = df.drop(columns=['non_empty_count'])
        return df.reset_index(drop=True)

    def procesar_archivos(self, archivos, directorio):
        # Limpia y guarda cada archivo en el directorio dado, eliminando la primera fila antes de salvar.
        for archivo in archivos:
            ruta_archivo = os.path.join(directorio, archivo)
            
            # Verifica si el archivo es válido antes de procesarlo.
            if not self.es_archivo_excel_valido(ruta_archivo):
                print(f'Error: El archivo "{archivo}" no es válido.')
                continue

            # Carga el libro de trabajo de Excel.
            book = load_workbook(ruta_archivo)

            # Procesa cada hoja del archivo.
            for num_hoja, hoja in enumerate(book.sheetnames, start=1):
                ws = book[hoja]
                data = ws.values  # Obtiene los valores de la hoja
                headers = next(data)  # Toma la primera fila como encabezados
                df = pd.DataFrame(data, columns=headers)  # Crea un DataFrame con los datos
                df_limpio = self.limpiar_hoja_actualizada(df)  # Limpia el DataFrame
                
                # Elimina la hoja original y crea una nueva hoja con los datos limpios.
                book.remove(ws)
                ws_nueva = book.create_sheet(title=hoja)
                
                # Inserta los datos limpios en la nueva hoja
                for r_idx, row in enumerate([df_limpio.columns.tolist()] + df_limpio.values.tolist(), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws_nueva.cell(row=r_idx, column=c_idx, value=value)
                
                # Comprueba si la primera fila solo tiene una celda con datos.
                primera_fila = list(ws_nueva.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                datos_en_primera_fila = sum(1 for celda in primera_fila if celda is not None)
                
                # Elimina la primera fila solo si contiene un único dato.
                if datos_en_primera_fila == 1:
                    ws_nueva.delete_rows(1)

            # Guarda los cambios en el archivo de Excel.
            book.save(ruta_archivo)
            print(f'Archivo guardado: {archivo}')


class AplicacionConsola:
    def __init__(self, in_colab, directorio_externo=None):
        """
        Inicializa la aplicación.
        
        Parámetros:
        - in_colab (bool): Indica si el entorno es Google Colab.
        - directorio_externo (str o None): Ruta de un directorio externo, si el programa es llamado desde otro.
        """
        self.gestor_directorios = GestorDirectorios(in_colab, directorio_externo)
        self.visualizador = VisualizadorDataframes()
        self.procesador = ProcesadorArchivosExcel()
        self.directorio = self.gestor_directorios.obtener_directorio_trabajo()

    def aviso_seguridad(self):
        """
        Muestra un mensaje de advertencia al usuario para que trabaje con archivos respaldados,
        ya que el programa puede modificar archivos originales.
        """
        print('\33[31m\n*** AVISO IMPORTANTE ***')
        print('Debe trabajar con archivos respaldados, ya que este programa puede modificar archivos originales.\33[37m')
        
        # Pide confirmación al usuario para continuar; si no confirma, el programa termina.
        if input('\n¿Desea continuar? (y/n): ').strip().lower() != 'y':
            print('Programa terminado para evitar modificaciones accidentales.')
            exit()

    def menu_principal(self):
        """
        Muestra el menú principal de la aplicación.
        """
        Utilidades.limpiar_consola()
        self.aviso_seguridad()  # Llama al aviso de seguridad antes de mostrar el menú principal
        
        while True:
            Utilidades.limpiar_consola()
            print('\33[32m\n--- MENÚ PRINCIPAL ---')
            print('1. Visualización de DataFrames')
            print('2. Limpieza de archivos')
            print('3. Salir\33[37m')
            opcion = input('Seleccione una opción: ').strip()
            
            if opcion == '1':
                # Opción 1: Listar archivos Excel en el directorio actual
                archivos_dict = UtilidadesArchivos.listar_archivos_xlsx(self.directorio)
                
                # Permitir al usuario seleccionar uno o más archivos para visualizar
                archivos = UtilidadesArchivos.seleccionar_archivos(archivos_dict)
                
                # Visualizar las hojas de los archivos seleccionados
                self.visualizador.visualizar_dataframes(self.directorio, archivos)
            
            elif opcion == '2':
                # Opción 2: Listar archivos Excel en el directorio actual
                archivos_dict = UtilidadesArchivos.listar_archivos_xlsx(self.directorio)
                
                # Permitir al usuario seleccionar uno o más archivos para limpieza
                archivos = UtilidadesArchivos.seleccionar_archivos(archivos_dict)
                
                # Procesar y limpiar los archivos seleccionados
                self.procesador.procesar_archivos(archivos, self.directorio)
            
            elif opcion == '3':
                # Opción 3: Salir del programa
                # Realiza una cuenta regresiva antes de salir y restaura el directorio original
                Utilidades.cuenta_regresiva(self.gestor_directorios)
                
                # Limpia la consola antes de finalizar el programa
                Utilidades.limpiar_consola()
                break
            
            else:
                # Manejo de entrada no válida por parte del usuario
                print('Opción no válida. Por favor, seleccione un número del menú.')


# Punto de entrada principal de la aplicación
if __name__ == '__main__':
    try:
        # Detectar si estamos en Google Colab
        import google.colab
        IN_COLAB = True
    except ImportError:
        # Si no es Colab, asumimos un entorno local (PC o Jupyter)
        IN_COLAB = False

    # Determinar si hay un directorio externo configurado por otro programa
    directorio_externo = os.getcwd()  # Usa el directorio actual como base si ya está definido

    # Inicializar la aplicación con el entorno y directorio externo
    app = AplicacionConsola(IN_COLAB, directorio_externo)
    app.menu_principal()
# %%